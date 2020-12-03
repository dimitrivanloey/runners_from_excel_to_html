import dominate
import openpyxl
import os
from dominate.tags import *

wb = openpyxl.load_workbook('kmp2.xlsx')
ws = wb.active
workbook_list = []
my_keys = []
teller = 1
# Creating a list with dictionaries for the rows
for col in range(0, ws.max_column):
    my_keys.append(ws.cell(row=1, column=col + 1).value)

for row in range(2, ws.max_row + 1):
    dictionary = {}
    for pos in range(2, len(my_keys) - 4):
        new_cell = "{:.2f}".format(ws.cell(row=row, column=pos + 1).value)
        print(new_cell)

        # dictionary[my_keys[pos]] = new_cell
    for pos in range(0, len(my_keys)):
        dictionary[my_keys[pos]] = ws.cell(row=row, column=pos + 1).value

    for key in dictionary:
        if key == 'Pos':
            dictionary[key] = teller
            teller += 1
        elif type(dictionary[key]) == float:
            dictionary[key] = "{:2.2f}".format(dictionary[key])
    workbook_list.append(dictionary)

# Round all the values to 2 decimals


# print(workbook_list)

# Converting the list to html table
doc = dominate.document(title="Excel spread sheet")
with doc.head:
    link(rel="stylesheet", href="style.css")
with doc:
    with div(id="horses").add(table()):
        with thead():
            dictionary = workbook_list[0]
            for key in dictionary.keys():
                table_header = td()
                table_header.add(key)

        for dictionary in workbook_list:
            table_row = tr(cls="excel_table_row")
            for key in dictionary:
                with table_row.add(td()):
                    p(dictionary[key])
print(str(doc))

# Test code
print(dictionary)
teller = 1
for key in dictionary:
    if key == 'Pos':
        dictionary[key] = teller
        teller += 1
    elif type(dictionary[key]) == float:
        dictionary[key] = "{:2.2f}".format(dictionary[key])
print(dictionary)
print(workbook_list)


file_path = os.path.abspath("new_webpage.html")
html_file = open(file_path, "w")
html_file.write(str(doc))
html_file.close()
